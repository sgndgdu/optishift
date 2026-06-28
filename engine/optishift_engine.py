#!/usr/bin/env python3
"""
OptiShift - Vardiya Optimizasyon Motoru
Google OR-Tools CP-SAT ile Adil Vardiya Planlama

Kurulum:
    pip install ortools openpyxl

Çalıştırma:
    python optishift_engine.py
"""

from ortools.sat.python import cp_model
import json

# ─── VERİ MODELİ ─────────────────────────────────────────────────────────────

STORE = {
    "store_id": "L-001",
    "store_name": "OptiShift Global",
    "connected_erp": "SAP_SuccessFactors",
}

# Çoklu Şube Destekli Personel Listesi
PERSONNEL = []

# Müsaitlik Haritası
AVAILABILITY = {}

# Kapasite Matrisi: {shift_idx: {day(0-6): exact_count}}
# Boşsa motor coverage-max moduna düşer
DEMAND_MATRIX = {}

# Tercih edilmeyen (sarı) günde çalıştırılan personelin puan telafi çarpanı
PREFERRED_NOT_MULTIPLIER = 1.5

# Her vardiyada en az 1 "primary" role_level personel olsun (soft constraint)
ENSURE_SENIOR_PER_SHIFT = False

# Ardışık maksimum çalışma günü (7 = devre dışı)
MAX_CONSECUTIVE_DAYS = 6

# Gece vardiyası (≥23:00 bitiş) sonrası sabah vardiyası (≤12:00 başlangıç) yasağı
NO_NIGHT_TO_MORNING = False

# ─── FABRİKA MODÜLÜ ───────────────────────────────────────────────────────────

# Ekip rotasyon kısıtı: {crew_id: shift_idx} — bu haftaki ekip-vardiya eşleşmesi
# Boşsa rotasyon kısıtı uygulanmaz
CREW_ROTATION = {}

# Personel→ekip haritası: {personnel_id: crew_id}
PERSONNEL_CREWS = {}

# Aynı ekip üyeleri kesinlikle aynı vardiyaya (True=hard, False=soft penalty)
CREW_SAME_SHIFT_HARD = False

# Fazla mesai eşiği: haftalık bu saatin üzeri = mesai (varsayılan 45 — haftada 45+ saat)
OVERTIME_THRESHOLD_HOURS = 45.0

# Yıllık maksimum fazla mesai saati (İş Kanunu 41. madde, varsayılan 270)
MAX_YTD_OVERTIME_HOURS = 270.0

# Adil mesai dağılımı: YTD mesai saati yüksek olanın ek vardiyası soft cezalandırılır
OVERTIME_FAIR_DISTRIBUTION = True


# Her bölgede günlük minimum çalışan kişi (Varsayılan şablon - Optimizasyon sırasında dinamik filtrelenecektir)
ZONE_DEMAND_PER_DAY = {
    "Kasa":   2,
    "Reyon":  1,
    "Teras":  1,
    "Mutfak": 1,
    "Resepsiyon": 1,
    "Kat Hizmetleri": 1,
    "Barista & Servis": 1,
}

RULES = {
    "max_weekly_hours": 45,
    "min_rest_hours":   11,
    "clopening_min_rest_hours": 13,
    # Adalet motoru v2 — burden multiplier'ları
    "weekend_multiplier":       1.2,
    "night_multiplier":         1.3,
    "preferred_not_multiplier": 1.5,
    "clopening_multiplier":     1.2,
    "hero_multiplier":          1.5,
}

DAYS = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]

SHIFTS = [
    {"name": "Sabah",  "start": "08:00", "end": "16:00", "base_points": 3},
    {"name": "Akşam",  "start": "16:00", "end": "24:00", "base_points": 5},
]

NUM_DAYS   = 7
NUM_SHIFTS = 2
SHIFT_HOURS = 8

# ─── YARDIMCI ────────────────────────────────────────────────────────────────

def _shift_minutes(shift: dict) -> tuple:
    """Vardiya start/end string'lerini toplam dakikaya çevirir. Gece geçişini destekler."""
    sh, sm = map(int, shift["start"].split(":"))
    eh, em = map(int, shift["end"].split(":"))
    start_min = sh * 60 + sm
    end_min   = eh * 60 + em
    if end_min <= start_min:   # 16:00–00:00 gibi gece geçişi
        end_min += 24 * 60
    return start_min, end_min

# ─── PUAN HESAPLAMA ──────────────────────────────────────────────────────────

def shift_points(day: int, shift_id: int) -> float:
    """
    Burden skoru: difficulty × duration_hours × weekend/night multiplier.
    Multiplier'lar RULES'tan okunur (müdür ayarlar).
    """
    shift = SHIFTS[shift_id] if shift_id < len(SHIFTS) else {}
    base = shift.get("base_points", 5)
    start_m, end_m = _shift_minutes(shift) if shift else (0, 480)
    hours = (end_m - start_m) / 60

    weekend_mult = RULES.get("weekend_multiplier", 1.2) if RULES.get("weekend_multiplier_enabled", True) else 1.0
    night_mult   = RULES.get("night_multiplier",   1.3) if RULES.get("night_multiplier_enabled",   True) else 1.0

    is_weekend = day in (5, 6)
    is_night   = shift.get("is_night", False)

    burden = base * hours
    if is_weekend: burden *= weekend_mult
    if is_night:   burden *= night_mult
    return round(burden)



def effective_points(person_id, day: int, shift_id: int) -> int:
    """Kişiye özel burden puanı: preferred_not günde çalışma telafi çarpanı uygulanır."""
    pts = shift_points(day, shift_id)
    if RULES.get("preferred_not_enabled", True) and get_avail(person_id, day) == "preferred_not":
        mult = RULES.get("preferred_not_multiplier", PREFERRED_NOT_MULTIPLIER)
        pts = int(pts * mult + 0.5)
    return pts


# ─── MODEL KURULUMU ───────────────────────────────────────────────────────────

def get_avail(person_id, d, _s=None):
    """Personelin gün d için müsaitlik durumunu döndürür.

    DB'den iki format gelebilir:
      - Basit string: "available" | "preferred_not" | "unavailable"
      - Dict (saat aralıklı): {"status": "preferred_not", "start": "09:00", "end": "17:00"}
    """
    avail_map = AVAILABILITY.get(person_id, {})
    # JSON serileştirme int key'leri string'e dönüştürür; ikisini de dene
    day_avail = avail_map.get(str(d)) if avail_map.get(str(d)) is not None else avail_map.get(d)
    if isinstance(day_avail, dict):
        return day_avail.get("status", "available")
    return day_avail if day_avail is not None else "available"

def build_model():
    model = cp_model.CpModel()
    num_p = len(PERSONNEL)

    # shifts[(p, d, s)] == 1 → personel p, gün d, vardiya s'de çalışıyor
    shifts = {
        (p, d, s): model.new_bool_var(f"shift_p{p}_d{d}_s{s}")
        for p in range(num_p)
        for d in range(NUM_DAYS)
        for s in range(NUM_SHIFTS)
    }

    # ── HARD CONSTRAINTS ─────────────────────────────────────────────────────

    # Aynı gün en fazla 1 vardiya
    for p in range(num_p):
        for d in range(NUM_DAYS):
            model.add_at_most_one(shifts[(p, d, s)] for s in range(NUM_SHIFTS))

    # Müsaitlik: "unavailable" günlerde kesinlikle çalışamaz (Kırmızı)
    for p_idx, person in enumerate(PERSONNEL):
        for d in range(NUM_DAYS):
            for s in range(NUM_SHIFTS):
                if get_avail(person["id"], d, s) == "unavailable":
                    model.add(shifts[(p_idx, d, s)] == 0)

    # Minimum dinlenme: Genel geçiş matrisi (N vardiya destekli)
    # Her (s1, s2) çifti için: ertesi güne geçen dinlenme süresi < min_rest_hours ise yasak.
    # Formül: rest_gap = (start_s2 + 1440) − end_s1
    # Örnek: Akşam 16-00 → Sabah 08-16: (480+1440)−1440 = 480 dk = 8s < 11s → YASAK ✓
    # Gece geçişli vardiyalar için _shift_minutes zaten end_min > 1440 döndürür.
    min_rest_min = RULES["min_rest_hours"] * 60
    forbidden_transitions = []
    for s1 in range(NUM_SHIFTS):
        _, end1 = _shift_minutes(SHIFTS[s1])
        for s2 in range(NUM_SHIFTS):
            start2, _ = _shift_minutes(SHIFTS[s2])
            rest_gap = (start2 + 1440) - end1
            if rest_gap < min_rest_min:
                forbidden_transitions.append((s1, s2))

    for p in range(num_p):
        for d in range(NUM_DAYS - 1):
            for s1, s2 in forbidden_transitions:
                model.add(shifts[(p, d, s1)] + shifts[(p, d + 1, s2)] <= 1)

    # Ardışık gün limiti: herhangi (MAX_CONSECUTIVE_DAYS+1) günlük pencerede en az 1 gün serbest
    if MAX_CONSECUTIVE_DAYS < NUM_DAYS:
        window = MAX_CONSECUTIVE_DAYS + 1
        for p in range(num_p):
            for start_d in range(NUM_DAYS - MAX_CONSECUTIVE_DAYS):
                model.add(
                    sum(shifts[(p, d, s)] for d in range(start_d, start_d + window) for s in range(NUM_SHIFTS))
                    <= MAX_CONSECUTIVE_DAYS
                )

    # Gececi→Sabahçı yasak: gece vardiyası (≥23:00 bitiş) → ertesi sabah (≤12:00 başlangıç)
    if NO_NIGHT_TO_MORNING:
        night_idxs = [s for s in range(NUM_SHIFTS) if _shift_minutes(SHIFTS[s])[1] >= 23 * 60]
        morning_idxs = [s for s in range(NUM_SHIFTS) if _shift_minutes(SHIFTS[s])[0] <= 12 * 60]
        for p in range(num_p):
            for d in range(NUM_DAYS - 1):
                for ns in night_idxs:
                    for ms in morning_idxs:
                        model.add(shifts[(p, d, ns)] + shifts[(p, d + 1, ms)] <= 1)

    # Haftalık maksimum saat limiti — gerçek vardiya sürelerini hesaba kat
    # Her vardiya tipinin gerçek süresini _shift_minutes ile hesapla (dakika cinsinden)
    shift_durations_min = []
    for s in range(NUM_SHIFTS):
        if s < len(SHIFTS):
            start_m, end_m = _shift_minutes(SHIFTS[s])
            shift_durations_min.append(end_m - start_m)
        else:
            shift_durations_min.append(SHIFT_HOURS * 60)

    max_weekly_minutes = RULES["max_weekly_hours"] * 60
    for p in range(num_p):
        # part-time personel için kendi max_weekly_hours'unu kullan
        person_max_min = int(PERSONNEL[p].get("max_weekly_hours", RULES["max_weekly_hours"])) * 60
        effective_max  = min(max_weekly_minutes, person_max_min)
        model.add(
            sum(
                shifts[(p, d, s)] * shift_durations_min[s]
                for d in range(NUM_DAYS)
                for s in range(NUM_SHIFTS)
            ) <= effective_max
        )

    # Bölge kotası: her gün yetkin kişi toplamı minimumu (min_per_day)
    # Sabah+akşam vardiyalarının toplamında o gün en az min_count kişi çalışmış olmalı
    for zone, min_count in ZONE_DEMAND_PER_DAY.items():
        skilled = [
            p_idx for p_idx, person in enumerate(PERSONNEL)
            if zone in person["skills"]
        ]
        for d in range(NUM_DAYS):
            if skilled:
                model.add(
                    sum(shifts[(p_idx, d, s)] for p_idx in skilled for s in range(NUM_SHIFTS))
                    >= min_count
                )

    # Demand-based exact coverage: DEMAND_MATRIX[shift_idx][day] = exact_count
    # Bu kısıt varsa o gün o vardiyaya tam N kişi atanır (ne fazla ne eksik)
    for s_idx, day_counts in DEMAND_MATRIX.items():
        for d, exact_count in day_counts.items():
            if exact_count > 0:
                model.add(
                    sum(shifts[(p, d, s_idx)] for p in range(num_p)) == exact_count
                )
            else:
                # 0 girildiyse o gün o vardiyaya kimse atanmaz
                model.add(
                    sum(shifts[(p, d, s_idx)] for p in range(num_p)) == 0
                )

    # ── FABRİKA: Ekip Rotasyon Kısıtı ───────────────────────────────────────
    # CREW_ROTATION: {crew_id: shift_idx} — bu haftaki ekip-vardiya ataması
    # Hard: ekip üyesi sadece atanan vardiyaya girer; diğer vardiyalara kesinlikle giremez.
    # Soft: atandığı vardiya dışına çıkarsa soft ceza uygulanır.
    if CREW_ROTATION:
        for p_idx, person in enumerate(PERSONNEL):
            pid = person["id"]
            crew = PERSONNEL_CREWS.get(pid)
            if crew and crew in CREW_ROTATION:
                assigned_shift = CREW_ROTATION[crew]
                if CREW_SAME_SHIFT_HARD:
                    # Hard: ekip vardiyası dışına çıkamaz
                    for s in range(NUM_SHIFTS):
                        if s != assigned_shift:
                            for d in range(NUM_DAYS):
                                model.add(shifts[(p_idx, d, s)] == 0)
                else:
                    # Soft: diğer vardiyaya atanırsa ×40 ceza (demand boşluğu varsa motor yine atayabilir)
                    for s in range(NUM_SHIFTS):
                        if s != assigned_shift:
                            for d in range(NUM_DAYS):
                                # shifts[(p_idx, d, s)] == 1 ise b == 1
                                b = model.new_bool_var(f"crew_viol_p{p_idx}_d{d}_s{s}")
                                model.add(shifts[(p_idx, d, s)] <= b)

    # ── FABRİKA: YTD Mesai Sınırı (Hard) ────────────────────────────────────
    # Yıllık limit dolmuş personelin haftası otomatik olarak normal eşiğe kısıtlanır.
    if MAX_YTD_OVERTIME_HOURS > 0:
        for p_idx, person in enumerate(PERSONNEL):
            ytd = float(person.get("ytd_overtime_hours", 0) or 0)
            remaining_ot_hours = MAX_YTD_OVERTIME_HOURS - ytd
            if remaining_ot_hours <= 0:
                # YTD dolmuş: bu hafta da normal eşiği aşamaz
                cap_minutes = int(OVERTIME_THRESHOLD_HOURS * 60)
                model.add(
                    sum(shifts[(p_idx, d, s)] * shift_durations_min[s]
                        for d in range(NUM_DAYS) for s in range(NUM_SHIFTS))
                    <= cap_minutes
                )

    # ── SOFT CONSTRAINTS (Ceza Değişkenleri) ─────────────────────────────────

    # preferred_not günlerde çalışmak istenmeyen ama zorunlu olabilir
    preferred_not_penalties = []
    if RULES.get("preferred_not_enabled", True):
        preferred_not_penalties = [
            shifts[(p_idx, d, s)]
            for p_idx, person in enumerate(PERSONNEL)
            for d in range(NUM_DAYS)
            for s in range(NUM_SHIFTS)
            if get_avail(person["id"], d, s) == "preferred_not"
        ]

    # Clopening (kapanış→açılış) soft cezası — OPTI-023.
    # Dinlenme yasal minimumun (min_rest_hours) üstünde ama clopening eşiğinin
    # altında kalan ardışık gün geçişleri yorucudur; motor alternatif varken kaçınır.
    clopening_min = int(RULES.get("clopening_min_rest_hours", 13)) * 60
    clopening_transitions = []
    clopening_penalties = []
    if RULES.get("clopening_enabled", True):
        for s1 in range(NUM_SHIFTS):
            _, end1 = _shift_minutes(SHIFTS[s1])
            for s2 in range(NUM_SHIFTS):
                start2, _ = _shift_minutes(SHIFTS[s2])
                rest_gap = (start2 + 1440) - end1
                if min_rest_min <= rest_gap < clopening_min:
                    clopening_transitions.append((s1, s2))
        for p in range(num_p):
            for d in range(NUM_DAYS - 1):
                for s1, s2 in clopening_transitions:
                    b = model.new_bool_var(f"clopening_p{p}_d{d}_s{s1}_{s2}")
                    # İki vardiya da atanmışsa b zorunlu 1; aksi halde minimize 0'a çeker
                    model.add(shifts[(p, d, s1)] + shifts[(p, d + 1, s2)] - 1 <= b)
                    clopening_penalties.append(b)

    # ── FABRİKA: Adil Mesai Dağılımı (Soft) ─────────────────────────────────
    # YTD mesai saati yüksek olanın overtime bölgesine girmesi soft cezalandırılır.
    # Eşiğin üzerindeki her olası gün için: ytd fazlaysa ceza ağırlığı artar.
    overtime_distribution_penalties = []
    if OVERTIME_FAIR_DISTRIBUTION and OVERTIME_THRESHOLD_HOURS > 0:
        ot_threshold_min = int(OVERTIME_THRESHOLD_HOURS * 60)
        for p_idx, person in enumerate(PERSONNEL):
            ytd = float(person.get("ytd_overtime_hours", 0) or 0)
            if ytd <= 0:
                continue
            # Bu kişi ne kadar çok YTD mesai yaptıysa, fazladan vardiyaları o kadar cezalandır.
            # Ceza ağırlığı: min(int(ytd / 10), 20) — makul bir skala
            penalty_weight = min(int(ytd / 10) + 1, 20)
            assigned_min = sum(
                shifts[(p_idx, d, s)] * shift_durations_min[s]
                for d in range(NUM_DAYS) for s in range(NUM_SHIFTS)
            )
            overtime_flag = model.new_bool_var(f"ot_flag_p{p_idx}")
            # overtime_flag == 1 ↔ bu hafta eşiği aşıyor
            # Soft: minimize eder, zorunlu olursa geçebilir
            model.add(assigned_min > ot_threshold_min).only_enforce_if(overtime_flag)
            model.add(assigned_min <= ot_threshold_min).only_enforce_if(overtime_flag.negated())
            overtime_distribution_penalties.append((overtime_flag, penalty_weight))

    # Minimum haftalık saat garantisi (soft) — part-time alt sınır.
    # Hard constraint yapılmaz: personel kırmızı günlerle haftayı kapatmışsa
    # plan kilitlenmesin diye eksik kalan dakikalar shortfall ile cezalandırılır.
    min_hours_shortfalls = []
    for p in range(num_p):
        min_minutes = int(PERSONNEL[p].get("min_weekly_hours", 0) or 0) * 60
        if min_minutes <= 0:
            continue
        assigned_minutes = sum(
            shifts[(p, d, s)] * shift_durations_min[s]
            for d in range(NUM_DAYS)
            for s in range(NUM_SHIFTS)
        )
        shortfall = model.new_int_var(0, min_minutes, f"min_hours_shortfall_p{p}")
        model.add(assigned_minutes + shortfall >= min_minutes)
        min_hours_shortfalls.append(shortfall)

    # ── ADİL PUAN OPTİMİZASYONU ──────────────────────────────────────────────

    person_scores = []
    weighted_scores = []  # Adaleti tartmak için part-time/full-time ağırlıklı skor

    for p_idx, person in enumerate(PERSONNEL):
        weekly_pts = sum(
            shifts[(p_idx, d, s)] * effective_points(person["id"], d, s)
            for d in range(NUM_DAYS)
            for s in range(NUM_SHIFTS)
        )
        total = model.new_int_var(0, 10000, f"total_score_p{p_idx}")
        # cumulative_burden = prev_score alanından geliyor (adalet motoru v2)
        model.add(total == int(person.get("cumulative_burden", person.get("prev_score", 0))) + weekly_pts)
        person_scores.append(total)

        # Part-time çalışanların hedeflenen saati daha düşük olduğu için, adalet skorlarını oranlıyoruz.
        # Çarpanlar: full_time = 1.0 (10), part_time = 0.6 (6)
        weight = int(RULES.get("part_time_weight_factor", 6)) if person.get("employment_type") == "part_time" else 10
        weighted = model.new_int_var(0, 5000, f"weighted_score_p{p_idx}")
        # weighted = (total * 10) / weight -> eğer part-time ise (total * 10) / 6, yani puanı suni olarak yüksek görünür, 
        # böylece algoritma ona daha fazla vardiya yazmak için yırtınmaz.
        model.add(weighted * weight == total * 10)
        weighted_scores.append(weighted)

    # Adalet farkı (Ağırlıklı): max_score - min_score → minimize
    max_score = model.new_int_var(0, 5000, "max_score")
    min_score = model.new_int_var(0, 5000, "min_score")
    model.add_max_equality(max_score, weighted_scores)
    model.add_min_equality(min_score, weighted_scores)
    fairness_gap = model.new_int_var(0, 5000, "fairness_gap")
    model.add(fairness_gap == max_score - min_score)

    # Toplam atama sayısı — yüksek olması isteniyor (coverage)
    num_p = len(PERSONNEL)
    max_possible = num_p * NUM_DAYS * NUM_SHIFTS
    total_assignments = model.new_int_var(0, max_possible, "total_assignments")
    model.add(total_assignments == sum(
        shifts[(p, d, s)]
        for p in range(num_p)
        for d in range(NUM_DAYS)
        for s in range(NUM_SHIFTS)
    ))

    # ── Kıdemli personel soft constraint ─────────────────────────────────────
    # Her vardiya × gün için primary personel yoksa ceza uygula (planı kilitlemez)
    senior_violation_penalties = []
    if ENSURE_SENIOR_PER_SHIFT and num_p > 0:
        primary_idxs = [
            p_idx for p_idx, person in enumerate(PERSONNEL)
            if person.get("role_level") == "primary"
        ]
        if primary_idxs:
            for s in range(NUM_SHIFTS):
                for d in range(NUM_DAYS):
                    # 1 if no primary assigned this shift/day, else 0
                    no_primary = model.new_bool_var(f"no_primary_s{s}_d{d}")
                    primary_sum = sum(shifts[(p, d, s)] for p in primary_idxs)
                    # no_primary == 1  ↔  primary_sum == 0
                    model.add(primary_sum == 0).only_enforce_if(no_primary)
                    model.add(primary_sum >= 1).only_enforce_if(no_primary.negated())
                    senior_violation_penalties.append(no_primary)

    # Amaç: adalet farkını minimize et (öncelik 1), coverage'ı maximize et (öncelik 2), preferred_not cezalarını minimize et (öncelik 3)
    # coverage terimi negatif → minimize ederken coverage artar
    # min_hours_shortfalls dakika cinsinden ×5: 8 saatlik açık ≈ 2400 ceza.
    # Bir vardiyalık fairness_gap artışı (~500-1700) bunun altında kalır,
    # yani motor alt sınırı doldurmayı adalet farkına tercih eder; yine de
    # kırmızı günler kapatmışsa plan kilitlenmez (soft).
    ot_penalty_term = sum(flag * w for flag, w in overtime_distribution_penalties) if overtime_distribution_penalties else 0

    model.minimize(
        fairness_gap * 100
        - total_assignments
        + sum(preferred_not_penalties) * 10
        + sum(senior_violation_penalties) * 50
        + sum(min_hours_shortfalls) * 5
        + sum(clopening_penalties) * int(RULES.get("clopening_penalty_weight", 30))
        + ot_penalty_term
    )

    return model, shifts, person_scores, fairness_gap


# ─── ÇÖZÜCÜ ──────────────────────────────────────────────────────────────────

def solve(silent: bool = False):
    model, shifts, person_scores, fairness_gap = build_model()

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 10.0
    solver.parameters.num_search_workers = 4
    solver.parameters.log_search_progress = False

    if not silent:
        print("OR-Tools CP-SAT çözücüsü çalışıyor...")
    status = solver.solve(model)

    status_map = {
        cp_model.OPTIMAL:   "OPTIMAL",
        cp_model.FEASIBLE:  "FEASIBLE (zaman sınırında en iyi bulunan)",
        cp_model.INFEASIBLE: "INFEASIBLE",
        cp_model.UNKNOWN:   "UNKNOWN",
    }
    if not silent:
        print(f"Çözücü durumu: {status_map.get(status, status)}")
        print(f"Çözüm süresi : {solver.wall_time:.3f}s")

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        print("\nHATA: Geçerli bir çözüm bulunamadı. Kısıtları gözden geçirin.")
        return None

    return solver, shifts, person_scores, fairness_gap


# ─── KONSOL ÇIKTISI ──────────────────────────────────────────────────────────

def print_schedule(solver, shifts, person_scores, fairness_gap):
    num_p = len(PERSONNEL)
    sep = "─" * 95

    print("\n" + "=" * 95)
    print(f"  {STORE['store_name']}  |  Haftalık Vardiya Planı")
    print("=" * 95)

    header = f"{'Personel / Yetenekler':<26}" + "".join(f"{d[:3]:^11}" for d in DAYS) + f"  {'Puan':>6}"
    print(header)
    print(sep)

    for p_idx, person in enumerate(PERSONNEL):
        skills_str = ",".join(person["skills"])
        label = f"{person['name']} [{skills_str}]"
        row = f"{label:<26}"
        weekly_pts = 0

        for d in range(NUM_DAYS):
            avail = AVAILABILITY[person["id"]].get(d, "available")
            cell = "—"
            for s in range(NUM_SHIFTS):
                if solver.value(shifts[(p_idx, d, s)]):
                    pts = shift_points(d, s)
                    weekly_pts += pts
                    tag = "S" if s == 0 else "A"
                    cell = f"{tag}·{pts}p"
                    if avail == "preferred_not":
                        cell += "!"   # tercih edilmeyen gün uyarısı
                    break
            if cell == "—" and avail == "unavailable":
                cell = "İZİN"
            row += f"{cell:^11}"

        total = solver.value(person_scores[p_idx])
        prev  = person["prev_score"]
        row  += f"  {total:>3}p (+{weekly_pts})"
        print(row)

    print(sep)
    gap = solver.value(fairness_gap)
    print(f"\nAdalet Farkı (max − min toplam puan): {gap} puan")
    print("Efsane: S=Sabah  A=Akşam  !=tercih edilmeyen gün  İZİN=unavailable\n")

    # Kural ihlali raporu
    print("── Haftalık Saat Kontrolleri " + "─" * 67)
    for p_idx, person in enumerate(PERSONNEL):
        n_shifts = sum(
            solver.value(shifts[(p_idx, d, s)])
            for d in range(NUM_DAYS) for s in range(NUM_SHIFTS)
        )
        hours = n_shifts * SHIFT_HOURS
        flag  = "✓" if hours <= RULES["max_weekly_hours"] else "⚠ FAZLA MESAİ!"
        print(f"  {person['name']:<22} {n_shifts} vardiya × {SHIFT_HOURS}s = {hours:>3}s   {flag}")

    print("\n── Bölge Kota Kontrolleri (min_per_day) " + "─" * 55)
    all_ok = True
    for zone, min_count in ZONE_DEMAND_PER_DAY.items():
        skilled = [
            p_idx for p_idx, person in enumerate(PERSONNEL)
            if zone in person["skills"]
        ]
        violations = []
        for d in range(NUM_DAYS):
            count = sum(
                solver.value(shifts[(p_idx, d, s)])
                for p_idx in skilled for s in range(NUM_SHIFTS)
            )
            if count < min_count:
                violations.append(f"{DAYS[d][:3]}({count}/{min_count})")
        if violations:
            all_ok = False
            print(f"  {zone:<10} ⚠ İhlal: {', '.join(violations)}")
        else:
            print(f"  {zone:<10} ✓ Kota sağlandı (günde min {min_count} kişi)")
    if all_ok:
        print("\n  Tüm bölge kotaları karşılandı.")
    print()


# ─── JSON EXPORT ─────────────────────────────────────────────────────────────

def export_json(solver, shifts, person_scores) -> dict:
    result = {
        "store_metadata": STORE,
        "generated_at": "2026-05-29",
        "week_schedule": [],
        "fairness_summary": [],
        "rules_applied": RULES,
    }

    for p_idx, person in enumerate(PERSONNEL):
        schedule = {"id": person["id"], "name": person["name"], "skills": person["skills"], "days": {}}
        for d in range(NUM_DAYS):
            assigned = None
            for s in range(NUM_SHIFTS):
                if solver.value(shifts[(p_idx, d, s)]):
                    assigned = {
                        "shift": SHIFTS[s]["name"],
                        "start": SHIFTS[s]["start"],
                        "end": SHIFTS[s]["end"],
                        "points": effective_points(person["id"], d, s),
                        "availability_status": AVAILABILITY[person["id"]].get(d, "available"),
                    }
                    break
            schedule["days"][DAYS[d]] = assigned or {"status": "off", "availability_status": AVAILABILITY[person["id"]].get(d, "available")}
        result["week_schedule"].append(schedule)

        result["fairness_summary"].append({
            "id": person["id"],
            "name": person["name"],
            "cumulative_burden": person.get("cumulative_burden", person.get("prev_score", 0)),
            "total_score": solver.value(person_scores[p_idx]),
        })

    path = "/Users/sefagundogdu/Desktop/OptiShift/optishift_result.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"  JSON → {path}")
    return result


# ─── EXCEL EXPORT ────────────────────────────────────────────────────────────

def export_excel(solver, shifts, person_scores):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  Excel export atlandı (pip install openpyxl)")
        return

    wb = openpyxl.Workbook()

    # Renk paletleri
    C_HEADER   = PatternFill("solid", fgColor="1F3864")
    C_MORNING  = PatternFill("solid", fgColor="DAEEF3")   # açık mavi
    C_EVENING  = PatternFill("solid", fgColor="FDE9D9")   # açık turuncu
    C_OFF      = PatternFill("solid", fgColor="F2F2F2")   # gri
    C_UNAVAIL  = PatternFill("solid", fgColor="FFD7D7")   # açık kırmızı
    C_PREF_NOT = PatternFill("solid", fgColor="FFFFE0")   # sarı uyarı
    C_OK       = PatternFill("solid", fgColor="C6EFCE")
    C_WARN     = PatternFill("solid", fgColor="FFC7CE")

    HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT       = Alignment(horizontal="left", vertical="center")

    # ── SEKMİ 1: Yönetici Özeti & KPI ────────────────────────────────────
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Yönetici Özeti & KPI"
    ws1.row_dimensions[1].height = 30
    ws1.row_dimensions[3].height = 22

    ws1.merge_cells("A1:F1")
    title_cell = ws1["A1"]
    title_cell.value = f"{STORE['store_name']}  |  Haftalık KPI Raporu"
    title_cell.font  = Font(bold=True, size=16, color="1F3864")
    title_cell.alignment = CENTER

    headers1 = ["Personel", "Yetenekler", "Vardiya Sayısı", "Toplam Saat", "Yasal Durum (≤45s)", "Adil Puan (Kümülatif)"]
    for col, h in enumerate(headers1, 1):
        c = ws1.cell(row=3, column=col, value=h)
        c.fill = C_HEADER
        c.font = HDR_FONT
        c.alignment = CENTER

    for p_idx, person in enumerate(PERSONNEL):
        row = 4 + p_idx
        n_shifts = sum(solver.value(shifts[(p_idx, d, s)]) for d in range(NUM_DAYS) for s in range(NUM_SHIFTS))
        hours    = n_shifts * SHIFT_HOURS
        score    = solver.value(person_scores[p_idx])
        ok       = hours <= RULES["max_weekly_hours"]

        ws1.cell(row=row, column=1, value=person["name"]).alignment = LEFT
        ws1.cell(row=row, column=2, value=", ".join(person["skills"])).alignment = LEFT
        ws1.cell(row=row, column=3, value=n_shifts).alignment = CENTER
        ws1.cell(row=row, column=4, value=hours).alignment = CENTER

        legal_cell = ws1.cell(row=row, column=5, value="✓ Uyumlu" if ok else "⚠ FAZLA MESAİ")
        legal_cell.fill      = C_OK if ok else C_WARN
        legal_cell.alignment = CENTER

        ws1.cell(row=row, column=6, value=score).alignment = CENTER

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 24
    for col in ["C", "D", "E", "F"]:
        ws1.column_dimensions[col].width = 22

    # Özet kutu
    summary_row = 4 + len(PERSONNEL) + 2
    ws1.cell(row=summary_row, column=1, value="Toplam Çalışan:").font = Font(bold=True)
    ws1.cell(row=summary_row, column=2, value=len(PERSONNEL))
    ws1.cell(row=summary_row + 1, column=1, value="Maks Haftalık Saat:").font = Font(bold=True)
    ws1.cell(row=summary_row + 1, column=2, value=RULES["max_weekly_hours"])
    ws1.cell(row=summary_row + 2, column=1, value="Min Dinlenme Süresi:").font = Font(bold=True)
    ws1.cell(row=summary_row + 2, column=2, value=f"{RULES['min_rest_hours']} saat")

    # ── SEKMİ 2: Haftalık Vardiya Planı ──────────────────────────────────
    ws2 = wb.create_sheet("Haftalık Vardiya Planı")
    ws2.row_dimensions[1].height = 24

    ws2.cell(row=1, column=1, value="Personel / Yetenekler").fill      = C_HEADER
    ws2.cell(row=1, column=1).font      = HDR_FONT
    ws2.cell(row=1, column=1).alignment = CENTER

    for d, day in enumerate(DAYS):
        c = ws2.cell(row=1, column=2 + d, value=day)
        c.fill      = C_HEADER
        c.font      = HDR_FONT
        c.alignment = CENTER

    total_col = 2 + NUM_DAYS
    c = ws2.cell(row=1, column=total_col, value="Haftalık Puan")
    c.fill      = C_HEADER
    c.font      = HDR_FONT
    c.alignment = CENTER

    for p_idx, person in enumerate(PERSONNEL):
        row = 2 + p_idx
        ws2.row_dimensions[row].height = 20

        name_label = f"{person['name']}\n[{', '.join(person['skills'])}]"
        nc = ws2.cell(row=row, column=1, value=name_label)
        nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        weekly_pts = 0
        for d in range(NUM_DAYS):
            avail  = AVAILABILITY[person["id"]].get(d, "available")
            cell   = ws2.cell(row=row, column=2 + d)
            cell.alignment = CENTER
            placed = False

            for s in range(NUM_SHIFTS):
                if solver.value(shifts[(p_idx, d, s)]):
                    pts = shift_points(d, s)
                    weekly_pts += pts
                    shift_name = "Sabah" if s == 0 else "Akşam"
                    warn = " !" if avail == "preferred_not" else ""
                    cell.value = f"{shift_name}\n({pts}p){warn}"
                    cell.fill  = C_MORNING if s == 0 else C_EVENING
                    placed = True
                    break

            if not placed:
                if avail == "unavailable":
                    cell.value = "İZİNLİ"
                    cell.fill  = C_UNAVAIL
                else:
                    cell.value = "—"
                    cell.fill  = C_OFF

        score_cell = ws2.cell(row=row, column=total_col)
        score_cell.value     = solver.value(person_scores[p_idx])
        score_cell.alignment = CENTER
        score_cell.font      = Font(bold=True)

    ws2.column_dimensions["A"].width = 26
    for i in range(NUM_DAYS):
        ws2.column_dimensions[get_column_letter(2 + i)].width = 14
    ws2.column_dimensions[get_column_letter(total_col)].width = 15

    # Renk lejantı
    legend_row = 2 + len(PERSONNEL) + 2
    ws2.cell(row=legend_row, column=1, value="Renk Lejantı:").font = Font(bold=True)
    legends = [
        (C_MORNING, "Sabah Açılış"),
        (C_EVENING, "Akşam Kapanış"),
        (C_UNAVAIL, "İzinli (Kırmızı)"),
        (C_PREF_NOT, "! = Tercih edilmeyen gün"),
    ]
    for i, (fill, label) in enumerate(legends):
        c = ws2.cell(row=legend_row + 1 + i, column=2, value=label)
        c.fill = fill

    path = "/Users/sefagundogdu/Desktop/OptiShift/optishift_schedule.xlsx"
    wb.save(path)
    print(f"  Excel → {path}")


# ─── ANA FONKSİYON ───────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  OptiShift - Vardiya Optimizasyon Motoru")
    print(f"  Mağaza   : {STORE['store_name']}")
    print(f"  Personel : {len(PERSONNEL)} kişi")
    print(f"  Dönem    : {NUM_DAYS} gün  |  {NUM_SHIFTS} vardiya tipi")
    print("=" * 60)

    result = solve()
    if result is None:
        return

    solver, shifts, person_scores, fairness_gap = result

    print_schedule(solver, shifts, person_scores, fairness_gap)

    print("── Dosya Çıktıları " + "─" * 77)
    export_json(solver, shifts, person_scores)
    export_excel(solver, shifts, person_scores)
    print()


def api_mode(payload: dict):
    """Next.js API route tarafından çağrılır. Dinamik JSON verisini kullanır."""
    import sys
    global PERSONNEL, AVAILABILITY, RULES, ZONE_DEMAND_PER_DAY, SHIFTS, NUM_SHIFTS, SHIFT_HOURS, DEMAND_MATRIX, ENSURE_SENIOR_PER_SHIFT, MAX_CONSECUTIVE_DAYS, NO_NIGHT_TO_MORNING, PREFERRED_NOT_MULTIPLIER
    global CREW_ROTATION, PERSONNEL_CREWS, CREW_SAME_SHIFT_HARD, OVERTIME_THRESHOLD_HOURS, MAX_YTD_OVERTIME_HOURS, OVERTIME_FAIR_DISTRIBUTION

    branch_id = payload.get("branchId", "L-001")
    ENSURE_SENIOR_PER_SHIFT = bool(payload.get("ensure_senior_per_shift", False))
    MAX_CONSECUTIVE_DAYS = int(payload.get("max_consecutive_days", 6))
    NO_NIGHT_TO_MORNING = bool(payload.get("no_night_to_morning", False))
    try:
        PREFERRED_NOT_MULTIPLIER = max(1.0, float(payload.get("preferred_not_multiplier", 1.5)))
    except (TypeError, ValueError):
        PREFERRED_NOT_MULTIPLIER = 1.5

    # Payload'dan gelen dinamik verileri global değişkenlere aktar
    raw_personnel    = payload.get("personnel", [])
    raw_availability = payload.get("availability", {})
    rules = payload.get("rules", {})

    if rules:
        RULES.update(rules)

    # Vardiya tanımlarını payload'dan oku (location shift_definitions)
    shifts_from_payload = payload.get("shifts")
    if shifts_from_payload and isinstance(shifts_from_payload, list) and len(shifts_from_payload) > 0:
        SHIFTS = [
            {
                "name":        s.get("name", f"Vardiya {i + 1}"),
                "start":       s.get("start", "08:00"),
                "end":         s.get("end",   "16:00"),
                "base_points": int(s.get("base_points", 5)),
            }
            for i, s in enumerate(shifts_from_payload)
        ]
        NUM_SHIFTS = len(SHIFTS)
        # Ortalama vardiya süresi → haftalık max vardiya sayısı hesabı için
        total_mins = sum((_shift_minutes(sh)[1] - _shift_minutes(sh)[0]) for sh in SHIFTS)
        SHIFT_HOURS = max(1, round((total_mins / NUM_SHIFTS) / 60))

    PERSONNEL = raw_personnel

    # Availability JSON keyleri string olabilir, integer'a çevirmemiz lazım (day_0 -> 0)
    AVAILABILITY = {}
    for p_id, av_dict in raw_availability.items():
        parsed = {}
        for day_str, status in av_dict.items():
            parsed[int(day_str)] = status
        AVAILABILITY[p_id] = parsed

    # ── Fabrika modülü parametreleri ─────────────────────────────────────────
    CREW_ROTATION    = payload.get("crew_rotation", {})      # {crew_id: shift_idx}
    PERSONNEL_CREWS  = payload.get("personnel_crews", {})    # {personnel_id: crew_id}
    CREW_SAME_SHIFT_HARD = bool(payload.get("crew_same_shift_hard", False))
    ot_rules         = payload.get("rules", {})
    OVERTIME_THRESHOLD_HOURS  = float(ot_rules.get("overtime_threshold_hours", 45.0))
    MAX_YTD_OVERTIME_HOURS    = float(ot_rules.get("max_ytd_overtime_hours", 270.0))
    OVERTIME_FAIR_DISTRIBUTION = bool(ot_rules.get("overtime_fair_distribution", True))
    # shift_idx dönüşümü: crew_rotation içindeki crew→shiftDefId'yi idx'e çevir
    if CREW_ROTATION and payload.get("shifts"):
        shifts_raw = payload.get("shifts") or []
        id_to_idx  = {}
        for i, s in enumerate(shifts_raw):
            if s.get("id"):   id_to_idx[str(s["id"])]   = i
            if s.get("name"): id_to_idx[str(s["name"])] = i
            id_to_idx[str(i)] = i
        CREW_ROTATION = {
            crew: id_to_idx.get(str(sid), 0)
            for crew, sid in CREW_ROTATION.items()
            if str(sid) in id_to_idx or any(str(i) == str(sid) for i in range(NUM_SHIFTS))
        }

    # Eğer o şube için hiç personel yoksa boş sonuç dön
    if not PERSONNEL:
        print(json.dumps({"error": f"Seçilen şubede ({branch_id}) aktif personel bulunamadı."}))
        return

    # Kapasite matrisi: {shiftDefId → {day → exact_count}} — demand-based scheduling
    # shiftDefId'yi shift index'ine çevir (payload'daki shifts sırası == index)
    DEMAND_MATRIX = {}
    raw_demand = payload.get("demand_matrix")
    if raw_demand and isinstance(raw_demand, dict):
        # shifts_from_payload varsa id→index map kur
        shifts_raw = payload.get("shifts") or []
        id_to_idx = {}
        for i, s in enumerate(shifts_raw):
            # Her shift için hem id hem name ile eşleşebilsin
            if s.get("id"):
                id_to_idx[str(s["id"])] = i
            if s.get("name"):
                id_to_idx[str(s["name"])] = i
            # Sayısal index fallback
            id_to_idx[str(i)] = i
        for shift_key, day_map in raw_demand.items():
            s_idx = id_to_idx.get(str(shift_key))
            if s_idx is None:
                # fallback: key sayısal ise doğrudan index olarak kullan
                try: s_idx = int(shift_key)
                except: continue
            if s_idx >= NUM_SHIFTS:
                continue
            parsed_days = {}
            for day_str, cnt in day_map.items():
                try:
                    parsed_days[int(day_str)] = int(cnt)
                except: pass
            if parsed_days:
                DEMAND_MATRIX[s_idx] = parsed_days

    # Bölge kotaları: payload'dan gel, yoksa global defaultlara dön
    # Format: {"Kasa": 2, "Reyon": 1} — skill adı → günlük minimum çalışan
    payload_quotas = payload.get("zone_quotas")
    all_branch_skills = set(skill for p in PERSONNEL for skill in p.get("skills", []))
    if payload_quotas and isinstance(payload_quotas, dict) and len(payload_quotas) > 0:
        # Sadece bu şubedeki personelin sahip olduğu yetenekleri içer
        ZONE_DEMAND_PER_DAY = {k: int(v) for k, v in payload_quotas.items() if k in all_branch_skills}
    else:
        # Payload'da kota yoksa global defaultları şube yeteneklerine göre daralt
        ZONE_DEMAND_PER_DAY = {k: v for k, v in ZONE_DEMAND_PER_DAY.items() if k in all_branch_skills}

    result = solve(silent=True)
    if result is None:
        print(json.dumps({"error": "Optimizasyon için uygun bir çözüm bulunamadı. Kısıtlamaları esnetmeyi deneyin."}))
        return

    solver, shifts, person_scores, fairness_gap = result

    output = {
        "fairness_gap": solver.value(fairness_gap),
        "assignments": [],
        "scores": {},
        "personnel": [],
        "senior_violations": [],  # vardiya × gün: primary personel atanamadı
        "overtime_summary": [],   # haftalık mesai özeti (fabrika modülü)
    }

    shift_durations_min = []
    for s in range(NUM_SHIFTS):
        if s < len(SHIFTS):
            start_m, end_m = _shift_minutes(SHIFTS[s])
            shift_durations_min.append(end_m - start_m)
        else:
            shift_durations_min.append(SHIFT_HOURS * 60)

    for p_idx, person in enumerate(PERSONNEL):
        output["scores"][person["id"]] = solver.value(person_scores[p_idx])
        output["personnel"].append({
            "id": person["id"],
            "name": person["name"],
            "skills": person.get("skills", []),
            "prev_score": person.get("prev_score", 0),
            "employment_type": person.get("employment_type", "full_time"),
            "availability": {str(k): v for k, v in AVAILABILITY.get(person["id"], {}).items()},
        })
        weekly_min = 0
        for d in range(NUM_DAYS):
            for s in range(NUM_SHIFTS):
                if solver.value(shifts[(p_idx, d, s)]):
                    dur = shift_durations_min[s] if s < len(shift_durations_min) else SHIFT_HOURS * 60
                    weekly_min += dur
                    output["assignments"].append({
                        "personnelId": person["id"],
                        "day":         d,
                        "shiftId":     s,
                        "start_time":  SHIFTS[s]["start"],
                        "end_time":    SHIFTS[s]["end"],
                        "points":      effective_points(person["id"], d, s),
                    })
        # Mesai özeti: eşiği aşan personeli raporla
        threshold_min = int(OVERTIME_THRESHOLD_HOURS * 60)
        if weekly_min > threshold_min:
            ot_hours = round((weekly_min - threshold_min) / 60, 2)
            ytd      = float(person.get("ytd_overtime_hours", 0) or 0)
            output["overtime_summary"].append({
                "personnelId":   person["id"],
                "name":          person["name"],
                "scheduled_hours": round(weekly_min / 60, 2),
                "overtime_hours":  ot_hours,
                "ytd_overtime_hours": ytd,
                "ytd_after":     round(ytd + ot_hours, 2),
                "at_limit":      (ytd + ot_hours) >= MAX_YTD_OVERTIME_HOURS,
            })

    # Kıdemli kısıt ihlallerini tespit et (primary atanamamış vardiya/gün kombinasyonları)
    if ENSURE_SENIOR_PER_SHIFT:
        primary_idxs = [
            p_idx for p_idx, person in enumerate(PERSONNEL)
            if person.get("role_level") == "primary"
        ]
        if primary_idxs:
            for s in range(NUM_SHIFTS):
                for d in range(NUM_DAYS):
                    has_primary = any(solver.value(shifts[(p, d, s)]) for p in primary_idxs)
                    has_anyone  = any(solver.value(shifts[(p, d, s)]) for p in range(len(PERSONNEL)))
                    if has_anyone and not has_primary:
                        output["senior_violations"].append({
                            "shift": SHIFTS[s]["name"],
                            "shift_idx": s,
                            "day": d,
                        })

    sys.stdout.write(json.dumps(output, ensure_ascii=False))


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "--api":
        import json as _json
        # stdin üzerinden tam JSON payload'ını oku
        raw_input = sys.stdin.read()
        if raw_input.strip():
            payload = _json.loads(raw_input)
            api_mode(payload)
        else:
            print(_json.dumps({"error": "Empty input received via stdin"}))
    else:
        # Manuel terminal testi için mock data oluştur veya hata ver
        print("Test modu (mock data) şu an API moduna çevrildiği için desteklenmiyor. --api argümanı ve stdin JSON payload'u gereklidir.")
